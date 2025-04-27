# inventory/views.py
import json

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404

from employees.models import AuditLog
from .forms import ProductForm, ProductItemForm, CategoryForm, BrandForm, CATEGORY_SPEC_FIELDS
from .models import ProductCategory, Brand, Product, ProductItem


def is_active_inventory_manager(user):
    if user.is_authenticated and (
            user.employee.role == 'Admin' or user.employee.role == 'Inventory Manager') and user.employee.is_active:
        return True
    return False


@login_required(login_url='login')
def inventoryManagerDashboard(request):
    context = {
        'logged_in_employee': request.user.employee
    }
    return render(request, 'inventoryManagerDashboard.html', context)


@login_required
def category_list(request):
    categories = ProductCategory.objects.all()
    return render(request, 'category_list.html',
                  {'categories': categories, 'logged_in_employee': request.user.employee})


@login_required
def category_create(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            AuditLog.objects.create(
                employee=request.user.employee,
                action=f'Created Category: {form.cleaned_data["name"]}')

            messages.success(request, 'Category created successfully.')
            return redirect('category_list')
    else:
        form = CategoryForm()
    return render(request, 'category_form.html', {'form': form, 'logged_in_employee': request.user.employee})


@login_required
def category_update(request, pk):
    category = get_object_or_404(ProductCategory, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()

            AuditLog.objects.create(
                employee=request.user.employee,
                action=f'Updated Category: {category.name}')

            messages.success(request, 'Category updated successfully.')
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'category_form.html', {'form': form, 'logged_in_employee': request.user.employee})


@login_required
def category_delete(request, pk):
    category = get_object_or_404(ProductCategory, pk=pk)
    if Product.objects.filter(category=category).exists():
        messages.error(request, 'Cannot delete category as it has associated products.')
        return redirect('category_list')
    category.delete()

    AuditLog.objects.create(
        employee=request.user.employee,
        action=f'Deleted Category: {category.name}'
    )

    messages.success(request, 'Category deleted successfully.')
    return redirect('category_list')


@login_required
def brand_delete(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    brand.delete()

    AuditLog.objects.create(
        employee=request.user.employee,
        action=f'Deleted Brand: {brand.name}'
    )

    messages.success(request, 'Brand deleted successfully.')
    return redirect('brand_list')


@login_required
def brand_create(request):
    if request.method == 'POST':
        form = BrandForm(request.POST)
        if form.is_valid():
            form.save()

            AuditLog.objects.create(
                employee=request.user.employee,
                action=f'Created Brand: {form.cleaned_data["name"]}'
            )

            messages.success(request, 'Brand created successfully.')
            return redirect('brand_list')
    else:
        form = BrandForm()
    return render(request, 'brand_form.html', {'form': form, 'logged_in_employee': request.user.employee})


@login_required
def brand_update(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        form = BrandForm(request.POST, instance=brand)
        if form.is_valid():
            form.save()

            AuditLog.objects.create(
                employee=request.user.employee,
                action=f'Updated Brand: {brand.name}'
            )
            messages.success(request, 'Brand updated successfully.')
            return redirect('brand_list')
    else:
        form = BrandForm(instance=brand)
    return render(request, 'brand_form.html', {'form': form, 'logged_in_employee': request.user.employee})


# inventory/views.py
@login_required
def brand_list(request):
    brands = Brand.objects.all()
    return render(request, 'brand_list.html', {'brands': brands, 'logged_in_employee': request.user.employee})


# inventory/views.py
@login_required
def product_list(request):
    products = Product.objects.all()
    for product in products:
        if product.category.name != "Accessories":
            computed_total = ProductItem.objects.filter(product=product, is_sold=False).count()
        else:
            computed_total = ProductItem.objects.filter(product=product).aggregate(total=Sum('quantity'))['total'] or 0
        # Bypass the property setter
        product.__dict__['total_stock'] = computed_total
    return render(request, 'product_list.html', {
        'products': products,
        'logged_in_employee': request.user.employee
    })

@login_required
def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save(commit=False)
            product.specifications = json.dumps(request.POST.get('specifications', {}))
            product.save()

            AuditLog.objects.create(
                employee=request.user.employee,
                action=f'Created Product: {product.name}'
            )

            messages.success(request, 'Product created successfully.')
            return redirect('stock_add', product_pk=product.pk)
    else:
        form = ProductForm()
    categories = ProductCategory.objects.all()
    brands = Brand.objects.all()
    return render(request, 'product_form.html', {
        'form': form,
        'categories': categories,
        'brands': brands,
        'logged_in_employee': request.user.employee
    })


@login_required
def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            product = form.save(commit=False)
            product.specifications = json.dumps(request.POST.get('specifications', {}))
            product.save()

            AuditLog.objects.create(
                employee=request.user.employee,
                action=f'Updated Product: {product.name}'
            )

            messages.success(request, 'Product updated successfully.')
            return redirect('product_list')
    else:
        form = ProductForm(instance=product)
    categories = ProductCategory.objects.all()
    brands = Brand.objects.all()
    return render(request, 'product_form.html', {
        'form': form,
        'categories': categories,
        'brands': brands,
        'logged_in_employee': request.user.employee
    })


@login_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    product.delete()

    AuditLog.objects.create(
        employee=request.user.employee,
        action=f'Deleted Product: {product.name}'
    )

    messages.success(request, 'Product deleted successfully.')
    return redirect('product_list')


@login_required
def product_stock(request, pk):
    product = get_object_or_404(Product, pk=pk)
    stock_items = ProductItem.objects.filter(product=product)

    if request.method == 'POST' and 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        category_fields = CATEGORY_SPEC_FIELDS.get(product.category.name, [])
        error_rows = []

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # Skip header row
            try:
                # Map row data to headers
                headers = [cell.value for cell in sheet[1]]
                data = {header: value for header, value in zip(headers, row)}

                # Determine item type
                item_type = ProductItem.SERIALIZED if data.get('IMEI') else ProductItem.NON_SERIAL

                # Validate serialized product IMEI
                if item_type == ProductItem.SERIALIZED and not data.get('IMEI'):
                    raise ValueError("Missing IMEI for serialized product.")

                # Validate prices
                cost_price = data.get('Cost Price')
                sale_price = data.get('Sale Price')
                if cost_price is None or sale_price is None or cost_price <= 0 or sale_price <= 0:
                    raise ValueError("Invalid cost or sale price.")

                # Combine specifications
                specifications = {
                    key: data[key] for key in category_fields if key in data and data[key] is not None
                }

                # Handle quantity for non-serial products
                quantity = data.get('Quantity', 1) if item_type == ProductItem.NON_SERIAL else 1

                # Create stock item
                ProductItem.objects.create(
                    product=product,
                    item_type=item_type,
                    serial_number=data.get('IMEI'),
                    quantity=quantity,
                    cost_price=cost_price,
                    sale_price=sale_price,
                    specifications=specifications
                )
            except Exception as e:
                error_rows.append(row_index)

        if error_rows:
            messages.warning(request,
                             f"Skipped rows: {', '.join(map(str, error_rows))} due to missing or invalid data.")
        else:
            messages.success(request, 'Stock items added successfully from Excel.')

        return redirect('product_stock', pk=product.pk)

    return render(request, 'product_stock.html', {
        'product': product,
        'stock_items': stock_items,
        'logged_in_employee': request.user.employee
    })


# inventory/views.py
@login_required
def stock_add(request, product_pk):
    product = get_object_or_404(Product, pk=product_pk)
    category = product.category.name  # Get the category name

    if request.method == 'POST':
        form = ProductItemForm(request.POST, request.FILES, category=category)
        if form.is_valid():
            stock_item = form.save(commit=False)
            stock_item.product = product

            # Explicitly set item_type based on category (belt-and-suspenders)
            if category == "Accessories":
                stock_item.item_type = ProductItem.NON_SERIAL
                # Ensure serial number is None for non-serialized
                stock_item.serial_number = None
            else:
                stock_item.item_type = ProductItem.SERIALIZED
                stock_item.quantity = 1

            specs_data = {}
            if category in CATEGORY_SPEC_FIELDS:
                for field_name in CATEGORY_SPEC_FIELDS[category]:
                    if field_name in form.cleaned_data:
                        specs_data[field_name] = form.cleaned_data[field_name]
            stock_item.specifications = specs_data

            stock_item.save()

            AuditLog.objects.create(
                employee=request.user.employee,
                action=f'Created stock: {stock_item.serial_number or stock_item.item_type} for product: {product.name} with specifications: {stock_item.specifications or "None"}'
            )

            messages.success(request, 'Stock item added successfully.')
            return redirect('product_stock', pk=product.pk)
        else:
            messages.error(request, 'Failed to add stock item. Please check the form.')

    else:  # GET request
        form = ProductItemForm(category=category)

    return render(request, 'stock_item_form.html', {
        'form': form,
        'product': product,
        'logged_in_employee': request.user.employee
    })


# inventory/views.py
@login_required
def stock_edit(request, pk):
    stock_item = get_object_or_404(ProductItem, pk=pk)
    category = stock_item.product.category.name  # Get the category name

    if request.method == 'POST':
        form = ProductItemForm(request.POST, request.FILES, instance=stock_item, category=category)
        if form.is_valid():
            stock_item = form.save(commit=False)
            stock_item.save()

            AuditLog.objects.create(
                employee=request.user.employee,
                action=f"Updated stock: {stock_item.serial_number or stock_item.item_type} for product: {stock_item.product.name} with specifications: {stock_item.specifications or 'None'}"
            )

            messages.success(request, "Stock item updated successfully.")
            return redirect("product_stock", pk=stock_item.product.pk)
    else:
        form = ProductItemForm(instance=stock_item, category=category)

    return render(request, "stock_item_form.html", {
        "form": form,
        "product": stock_item.product,
        "logged_in_employee": request.user.employee
    })


@login_required
def stock_delete(request, pk):
    stock_item = get_object_or_404(ProductItem, pk=pk)
    stock_item.delete()

    AuditLog.objects.create(
        employee=request.user.employee,
        action=f'Deleted stock: {stock_item.serial_number or stock_item.item_type} for product: {stock_item.product.name} with specifications: {stock_item.specifications or "None"}'
    )

    messages.success(request, 'Stock item deleted successfully.')
    return redirect('product_stock', pk=stock_item.product.pk)


@login_required
def download_excel_template(request, pk):
    product = get_object_or_404(Product, pk=pk)
    category_fields = CATEGORY_SPEC_FIELDS.get(product.category.name, [])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={product.name}_template.xlsx'

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = f'{product.name} Template'

    # Add headers
    headers = ['Cost Price', 'Sale Price'] + category_fields
    if product.category.name != "Accessories":  # Include IMEI only for serialized categories
        headers.insert(0, 'IMEI')
    else:  # Include Quantity for non-serial categories
        headers.append('Quantity')
    sheet.append(headers)

    # Example row
    example_row = [100.00, 150.00] + ['Example' for _ in category_fields]
    if product.category.name != "Accessories":
        example_row.insert(0, 'SN12345')
    else:
        example_row.append(10)  # Example quantity for non-serial products
    sheet.append(example_row)

    wb.save(response)
    return response


def get_brands_by_category(request):
    category_id = request.GET.get('category_id')
    if not category_id:
        return JsonResponse({'error': 'Category ID is required'}, status=400)

    brands = Brand.objects.filter(categories__id=category_id).values('id', 'name')
    return JsonResponse(list(brands), safe=False)
